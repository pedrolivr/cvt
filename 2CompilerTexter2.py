import cv2
import pytesseract
import os
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook

# Configurar o caminho do executável do Tesseract, se necessário
pytesseract.pytesseract.tesseract_cmd = r'C:/Users/pe.oliveira/AppData/Local/Programs/Tesseract-OCR/tesseract.exe'

# Função para ler texto de uma região específica da imagem
def read_text_from_region(image_path, region):
    """
    Lê texto de uma região específica de uma imagem.

    Args:
    image_path (str): Caminho para a imagem.
    region (tuple): Região da imagem no formato (x, y, largura, altura).

    Returns:
    str: Texto lido da região.
    """
    # Carregar a imagem
    image = cv2.imread(image_path)
    
    # Extrair a região de interesse (ROI)
    x, y, w, h = region
    roi = image[y:y+h, x:x+w]
    
    # Converter a ROI para escala de cinza
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    
    # Aplicar OCR na ROI usando o idioma português
    text = pytesseract.image_to_string(gray, lang='por')
    
    return text

# Função para selecionar o diretório
def select_directory():
    root = tk.Tk()
    root.withdraw()  # Ocultar a janela principal
    folder_selected = filedialog.askdirectory()
    return folder_selected

# Selecionar o diretório contendo as imagens
directory = select_directory()

# Especificar o caminho do template do Excel
template_path = r'C:/Users/pe.oliveira/OneDrive - EGIS Group/Bureau/cvt/MODELO-RESPOSTAS-CARIMBO_R01.xlsx'

# Definir as regiões de interesse (ROIs)
regions = [
    (3820, 2960, 900, 150),
    (3720, 3100, 300, 70),
    (4200, 3100, 500, 70),
    (3720, 3200, 900, 70),
    (3720, 3250, 300, 70),
    (4200, 3250, 300, 70),
    (4, 3260, 150, 50),
    (4330, 2570, 400, 50),
    (3720, 2700, 500, 150),
    (4225, 2700, 500, 150),
    (4300, 2875, 250, 70)
]

# Definir os títulos para cada região
region_titles = [

    "TÍTULO CARIMBO",
    "RODOVIA CARIMBO",
    "TRECHO CARIMBO",
    "CÓDIGO CARIMBO",
    "ESCALAS",
    "FOLHA CARIMBO",
    "REVISÃO INTERNA CARIMBO",
    "RESPONSÁVEL TÉCNICO CARIMBO",
    "LOGO ANTT CARIMBO",
    "LOGO CLIENTE CARIMBO",
    "LOGO EGIS CARIMBO"
]

# Verificar se o número de títulos corresponde ao número de regiões
if len(region_titles) != len(regions):
    raise ValueError("O número de títulos deve corresponder ao número de regiões.")

# Lista para armazenar os dados para o Excel
excel_data = []

# Obter o número total de imagens a serem processadas
total_images = len([filename for filename in os.listdir(directory) if filename.lower().endswith(('.png', '.jpg', '.jpeg'))])

# Criar uma barra de progresso única para todas as imagens
with tqdm(total=total_images, desc="Processando imagens", unit="ARQUIVO") as pbar:
    # Processar todas as imagens no diretório selecionado
    for idx, filename in enumerate(os.listdir(directory)):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
            image_path = os.path.join(directory, filename)
            output_file = os.path.join(directory, f'resultados_{os.path.splitext(filename)[0]}.txt')
            
            with open(output_file, 'w', encoding='utf-8') as file:
                file.write(f'Resultados para {filename}:\n')
                # Armazenar apenas os primeiros 35 caracteres do nome do arquivo
                image_data = {"ID": idx + 1, "ARQUIVO": filename[:35]}
                for i, region in enumerate(regions):
                    text = read_text_from_region(image_path, region)
                    title = region_titles[i]
                    file.write(f'Texto da {title}:\n')
                    file.write(text)
                    file.write('\n' + '-' * 30 + '\n')
                    image_data[title] = text
                file.write('\n' + '=' * 30 + '\n')
                excel_data.append(image_data)
            
            # Atualizar a barra de progresso
            pbar.update(1)

# Carregar o template do Excel
workbook = load_workbook(template_path)
sheet = workbook.active

# Inserir os cabeçalhos no template
sheet.cell(row=1, column=1, value="ID")
sheet.cell(row=1, column=2, value="ARQUIVO")
for col, title in enumerate(region_titles, start=3):
    sheet.cell(row=1, column=col, value=title)

# Inserir os dados processados no template
row = 2  # Assumindo que a primeira linha é o cabeçalho
for data in excel_data:
    sheet.cell(row=row, column=1, value=data["ID"])  # Índice
    sheet.cell(row=row, column=2, value=data["ARQUIVO"])  # Nome da imagem
    for col, title in enumerate(region_titles, start=3):
        sheet.cell(row=row, column=col, value=data[title])  # Texto das regiões
    row += 1

# Salvar o arquivo Excel com os dados inseridos
excel_output_file = os.path.join(directory, 'resultados_com_template.xlsx')
workbook.save(excel_output_file)

print(f"Processamento concluído. Resultados exportados para arquivos de texto separados e para {excel_output_file}")
